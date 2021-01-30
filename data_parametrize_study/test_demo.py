# Time: 2021/1/27 17:07
# Author: jiangzhw
# FileName: test_demo.py
import json

import openpyxl
import pytest
import yaml
import csv
import pandas as pd


class TestDemo:

    @pytest.mark.parametrize("userid,name,email",
                             yaml.load(open(r"D:\PythonPro\MyProjects\data_parametrize_study\data.yml",
                                            encoding="utf_8"))["demo"])
    def test_yaml1(self, userid, name, email):
        print(f"\nid:{userid}\nname:{name}\nemail:{email}")

    def test_yaml2(self):
        with open(r"D:\PythonPro\MyProjects\data_parametrize_study\data.yml",
                  encoding="utf_8") as f:
            data = yaml.load(f)
        # print(data)
        # print(data["demo"])
        # print(data["demo"][0])
        # print(data["demo"][0][0])
        print(data)
        print(data["login"])
        print(data["login"][1])
        print(data["message"][1])
        print(data["message"][0][2])

    def test_csv1(self):
        result = []
        with open('data.csv', 'r') as f:
            reader = csv.reader(f)
            # 获取第一行数据
            # result = list(reader)
            # print(result[0])
            # 获取第一列数据
            for row in reader:
                print(row)
                print(row[0])

    def test_csv2(self):
        with open('data.csv', 'r') as f:
            csv_data = pd.read_csv(f)
            # print(type(csv_data))
            print(csv_data["name"])
            print(csv_data["name"][0])
            # print(csv_data.columns.tolist())

    def test_excel(self):
        # xlrd/xlwt 或openpyxl或pandas
        # TODO：pandas读取xlsx文件报错问题待解决
        # with open('data.xlsx', 'r', encoding="utf-8") as f:
        #     excel_data = pd.read_excel(f)
        #     print(excel_data.head())

        workbook = openpyxl.load_workbook('data.xlsx')
        worksheet = workbook.get_sheet_by_name('Sheet1')
        # print(list(worksheet.rows))
        row2 = [item.value for item in list(worksheet.rows)[1]]
        print('第3行值', row2)
        col1 = [item.value for item in list(worksheet.columns)[1]]
        print('第2列值', col1)
        cell_2_3 = worksheet.cell(row=1, column=2).value
        print('第1行第2列值', cell_2_3)
        max_row = worksheet.max_row
        print('最大行', max_row)

    # json 四种方法：
    # loads：将string转换为dict
    # dumps：将dict转换为string
    # load：将json格式字符串转换为dict
    # dump：将dict转换为json格式字符串
    def test_json1(self):
        with open(r"D:\PythonPro\MyProjects\data_parametrize_study\data.json",
                  "r", encoding="utf-8") as f:
            data = json.load(f)
            print(data)
            # print(data[0])
            print(data["user1"])

    def test_txt1(self):
        with open(r"D:\PythonPro\MyProjects\data_parametrize_study\data.txt",
                  "r", encoding="utf-8") as f:
            data = pd.read_table(f, header=None, sep=" ")
            # data = data.head(1).dropna(inplace=True)
            print(data.drop(index=[0]))
            print(type(data))
            print(data)
            # print(data.head(1))
            # print(data[1])
            # 测试

